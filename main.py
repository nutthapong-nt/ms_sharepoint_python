
from json import load
from mysql.connector import connect
from time import sleep

config=load(open(r".\config.json"))

def pull_data_from_sharepoint(username,password):
    
    from office365.runtime.auth.authentication_context import AuthenticationContext
    from office365.sharepoint.client_context import ClientContext
    from office365.sharepoint.files.file import File
    print("login sharepoint...")
    sharepoint_URL=config["sharepoint_URL"]
    sharepoint_Path=config["sharepoint_Path"]
    ctx_auth = AuthenticationContext(sharepoint_URL)
    ctx_auth.acquire_token_for_user(username, password)  
    print("find the file...")
    ctx = ClientContext(sharepoint_URL, ctx_auth)
    response = File.open_binary(ctx, sharepoint_Path)
    print("download the file...")
    with open(config["tempxlsx_path"], "wb") as local_file:
        local_file.write(response.content)

def update_data():
    
    from openpyxl import load_workbook
    print("prepare data...")

    mainData=dict()
    SubData=dict()
 
    wb_obj = load_workbook(config["tempxlsx_path"])
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    
    for i in range(2,m_row+1):

        value1 = sheet_obj.cell(row = i, column = 1).value
        value2 = sheet_obj.cell(row = i, column = 2).value
        value3 = sheet_obj.cell(row = i, column = 3).value
        value4 = sheet_obj.cell(row = i, column = 4).value
        value5 = sheet_obj.cell(row = i, column = 5).value
        
        if value1 not in mainData:
            mainData[value1]=value2
        
        SubData[value3]=(value4,value5,value1)
        
    def WriteDB(query):
        dbConfig = config["database"]
        conn = connect(**dbConfig)
        cursor = conn.cursor()
        cursor.execute(query)
        conn.commit()
        conn.close()

    #CLEAR OLD DATA
    print("clear old data...")
    for query in ["DELETE FROM `...` WHERE 1","DELETE FROM `...` WHERE 1"]:
        WriteDB(query)
    
    #INSERT NEW DATA
    print("insert new data...")
    #INSERT MAIN DATA
    query = "INSERT INTO `...`(`...`, `...`) VALUES "+\
        ",".join(f"('{i}','{mainData[i]}')" for i in mainData)
    WriteDB(query)
    
    #insert SubData
    query = "INSERT INTO `...`(`...`, `...`, `...`, `...`) VALUES "+\
        ",".join(f"('{i}','{SubData[i][0]}','{SubData[i][2]}','{SubData[i][1]}')" for i in SubData)
    WriteDB(query)

def main():
    
    import rsa
    
    f=open(r".\userconfig.py","r")

    enusername,enpassword=eval(f.read())
    f.close()

    key = rsa.PrivateKey("*secret*")
    
    username = rsa.decrypt(enusername, key).decode()
    password = rsa.decrypt(enpassword, key).decode()
    
    pull_data_from_sharepoint(username, password)
    update_data()
    print("complete...")
    confirm=input(sleep(30))
    

main()